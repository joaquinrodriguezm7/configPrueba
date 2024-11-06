from openpyxl import Workbook
from openpyxl.styles import Protection, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import xlwings as xw
from win32com.client import Dispatch
import os

def leer_configuracion(archivo_config):
    config = {}
    with open(archivo_config, 'r') as f:
        for linea in f:
            if '=' in linea:
                clave, valor = linea.strip().split('=')
                config[clave.strip()] = valor.strip().split(' | ')
    return config

def inicializar_hoja(ws, archivo_config):
    # Leer el archivo de configuración
    config = leer_configuracion(archivo_config)
    
    # Obtener los nombres de los alumnos y preguntas desde la configuración
    alumnos = config.get('alumnos', [])
    preguntas = config.get('preguntas', [])
    
    # Escribir alumnos en las celdas A2:A11 (número y nombre), B2:B11 (RUT), y C2:C11 (ID) y protegerlas
    for i in range(2, 2 + len(alumnos)):
        datos_alumno = alumnos[i-2].split(' ')  # Dividir por espacios para obtener las partes del alumno
        
        # Escribir número de lista y nombre en columna A
        ws[f'A{i}'].value = f'{datos_alumno[0]} {datos_alumno[1]}'
        ws[f'A{i}'].protection = Protection(locked=True)
        
        # Escribir RUT en columna B
        ws[f'B{i}'].value = datos_alumno[2]
        ws[f'B{i}'].protection = Protection(locked=True)
        
        # Escribir Identificador en columna C
        ws[f'C{i}'].value = datos_alumno[3]
        ws[f'C{i}'].protection = Protection(locked=True)

    # Escribir preguntas en las celdas D1 hasta K1 (dependiendo del número de preguntas) y protegerlas
    for j in range(4, 4 + len(preguntas)):  # Columnas D a K (columnas 4 a 11)
        ws.cell(row=1, column=j).value = preguntas[j-4]
        ws.cell(row=1, column=j).protection = Protection(locked=True)


    # Crear lógica para preguntas que acepten enteros y decimales, incluyendo negativos (P1)
    rng_enteros = config.get('rng_enteros', [])
    
    dv_enteros = DataValidation(type="decimal",
                        operator="between",
                        formula1=None,
                        formula2=None,
                        showErrorMessage=True,
                        errorTitle="Entrada inválida",
                        error="Solo se permiten números enteros o decimales con separador ','")

    for rango in rng_enteros:
        for row in ws[rango]:
            for cell in row:
                dv_enteros.add(cell)

    # Crear lista desplegable para preguntas abiertas simples (P2)
    rng_abierta_simple = config.get('rng_abierta_simple', [])
    
    dv_abierta_simple = DataValidation(type="list",
                        formula1='"0,1,2"',
                        showErrorMessage=True,
                        errorTitle="Valor Inválido",
                        error="El valor debe ser uno de los seleccionados en la lista.")

    for rango in rng_abierta_simple:
        for row in ws[rango]:
            for cell in row:
                dv_abierta_simple.add(cell)

    # Crear lógica para preguntas que acepten fracciones (P3)
    rng_fracciones = config.get('rng_fracciones', [])

# Crear la validación de datos para las fracciones
    dv_fracciones = DataValidation(
        type="custom",
        formula1='=AND(ISNUMBER(VALUE(LEFT(A1,FIND("/",A1)-1))),ISNUMBER(VALUE(MID(A1,FIND("/",A1)+1,LEN(A1)-FIND("/",A1)))),COUNTIF(A1,"*/?*")=1)',
        showErrorMessage=True,
        error="Solo se permiten fracciones en formato numerador/denominador, ej: 1/2",
        errorTitle="Entrada inválida"
    )

# Agregar la validación de datos al worksheet (fuera del bucle)
    ws.add_data_validation(dv_fracciones)

# Iterar sobre los rangos y aplicar la validación a las celdas
    for rango in rng_fracciones:
        for row in ws[rango]:
            for cell in row:
            # Ajustar la fórmula para cada celda utilizando la referencia de la celda actual
                dv_fracciones.formula1 = dv_fracciones.formula1.replace('A1', cell.coordinate)
                dv_fracciones.add(cell)
                cell.number_format = '@'  # Formato de texto
                


    # Crear lista desplegable para preguntas de selección única (P4)
    rng_seleccion_unica = config.get('rng_seleccion_unica', [])
    
    dv_seleccion_unica = DataValidation(type="list",
                        formula1='"A, B, C, D, N, N/C"',
                        showErrorMessage=True,
                        errorTitle="Valor Inválido",
                        error="El valor debe ser uno de los seleccionados en la lista.")

    for rango in rng_seleccion_unica:
        for row in ws[rango]:
            for cell in row:
                dv_seleccion_unica.add(cell)

    # Crear lógica para preguntas que acepten pares ordenados (P6)
    rng_par_ordenado = config.get('rng_par_ordenado', [])

# Crear la validación de datos, manteniendo la referencia genérica 'A1'
    dv_par_ordenado = DataValidation(
        type="custom",
        formula1='=AND(ISNUMBER(VALUE(LEFT(A1,FIND(";",A1)-1))),ISNUMBER(VALUE(MID(A1,FIND(";",A1)+1,LEN(A1)-FIND(";",A1)))),COUNTIF(A1,"*;?*")=1)',
        showErrorMessage=True,
        error="Solo se permiten valores en formato de par ordenado, ej: X;Y",
        errorTitle="Entrada inválida"
    )

# Agregar la validación de datos al worksheet (fuera del bucle)
    ws.add_data_validation(dv_par_ordenado)

# Iterar sobre los rangos y aplicar la validación a las celdas
    for rango in rng_par_ordenado:
        for row in ws[rango]:
            for cell in row:
            # Ajustar la fórmula para cada celda utilizando la referencia de la celda actual
            # Nota: La fórmula se mantiene genérica y no cambia durante la iteración
                dv_par_ordenado.formula1 = dv_par_ordenado.formula1.replace('A1', cell.coordinate)
                dv_par_ordenado.add(cell)
                cell.number_format = '@'  # Formato de texto

    # Agregar validaciones a la hoja
    ws.add_data_validation(dv_enteros)
    ws.add_data_validation(dv_abierta_simple)
    ws.add_data_validation(dv_seleccion_unica)

    # Desbloquear solo las celdas B2:H11 (Preguntas 1 a 6)
    for row in ws_preguntas['B2:H11']:
        for cell in row:
            cell.protection = Protection(locked=False)

    # Bloquear las demás celdas por defecto
    ws.protection.sheet = False 

# Crear un archivo Excel
wb = Workbook()

# Inicializar la hoja principal
ws_preguntas = wb.active
ws_preguntas.title = 'Preguntas'
inicializar_hoja(ws_preguntas, r"C:\Users\joaquin.rodriguezm\Desktop\config.txt")

# Asignar estilos a la hoja de preguntas

start_row = 1
end_row = 11
start_column = 1  # Columna A
end_column = 9    # Columna H

sheet = ws_preguntas

# Crear un estilo de borde
thin = Side(border_style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Aplicar el borde a cada celda del rango
for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_column, max_col=end_column):
    for cell in row:
        cell.border = border

ws_preguntas.column_dimensions['H'].width = 40
# Crear e inicializar una nueva hoja
ws_respuestas = wb.create_sheet(title='Respuestas') 
inicializar_hoja(ws_respuestas, r"C:\Users\joaquin.rodriguezm\Desktop\config.txt")
#ws_respuestas.protection.sheet = True

# Referenciar los datos de B2:E11 en la hoja 'Respuestas', pero dejando vacía si la celda original está vacía
for i in range(2, 12):
    for j in range(4, 8):
        cell_respuesta = ws_respuestas.cell(row=i, column=j)
        cell_pregunta = ws_preguntas.cell(row=i, column=j)
        cell_respuesta.value = f'=IF(Preguntas!{cell_pregunta.coordinate}="","",Preguntas!{cell_pregunta.coordinate})'

# Referenciar los datos de la Pregunta 6 (columna G) en la hoja 'Respuestas'
for i in range(2, 12):
    cell_respuesta = ws_respuestas.cell(row=i, column=9)
    cell_pregunta = ws_preguntas.cell(row=i, column=9)
    cell_respuesta.value = f'=IF(Preguntas!{cell_pregunta.coordinate}="","", "("&Preguntas!{cell_pregunta.coordinate}&")")'

# Referenciar los datos de K1:K10 en la hoja 'Datos' a F2:F11 en la hoja 'Respuestas'
for i in range(2, 12):
    cell_respuesta = ws_respuestas.cell(row=i, column=8) 
    cell_dato = ws_preguntas.cell(row=i-1, column=11)  
    cell_respuesta.value = f"=Datos!K{i-1}"


# Crear la hoja de Datos
ws_datos = wb.create_sheet(title='Datos')
ws_datos.sheet_state = 'veryHidden'

# Guardar el archivo Excel
file_path = r'C:\Users\joaquin.rodriguezm\Desktop\cfgPrueba.xlsx'
wb.save(file_path)

# Usar xlwings para agregar el código VBA que genera los CheckBoxes
app = xw.App(visible=False) 
wb = app.books.open(file_path)

# Guardar y cerrar el archivo
wb.save()
wb.close()
app.quit()

def cambio_formato(ruta_archivo):
    excel = Dispatch('Excel.Application')
    excel.Visible = False
    libro = excel.Workbooks.Open(ruta_archivo)

    ruta_archivo_xlsm = ruta_archivo.replace('.xlsx', '.xlsm')
    libro.SaveAs(ruta_archivo_xlsm, FileFormat=52)
    libro.Close(SaveChanges = True)
    excel.Quit()

ruta_archivo = r'C:\Users\joaquin.rodriguezm\Desktop\cfgPrueba.xlsx'
cambio_formato(ruta_archivo)
os.remove(ruta_archivo)

import win32com.client

# Ruta al archivo de Excel
ruta_archivo = r'C:\Users\joaquin.rodriguezm\Desktop\cfgPrueba.xlsm'

# Crear un objeto de Excel
excel = win32com.client.Dispatch("Excel.Application")

# Abrir el archivo
libro = excel.Workbooks.Open(ruta_archivo)

# Código de la macro que deseas agregar
codigo_checkboxes = '''
Sub AddCheckBoxes()
    Dim ws As Worksheet
    Dim wsLinks As Worksheet
    Set ws = ThisWorkbook.Sheets("Preguntas")
    Set wsLinks = ThisWorkbook.Sheets("Datos")
    
    Dim i As Integer
    Dim j As Integer
    Dim leftPos As Double
    Dim topPos As Double
    Dim checkBoxWidth As Double
    Dim labels As Variant
    checkBoxWidth = 45  ' Aumentar el ancho de espaciado
    
    ' Etiquetas para los checkboxes
    labels = Array("OP1", "OP2", "OP3", "OP4", "OP5")
    
    ' Crear los checkboxes en el rango H2:H11 
    For i = 2 To 11
        topPos = ws.Cells(i, 8).Top  ' Cambiado a columna H (8)
        leftPos = ws.Cells(i, 8).Left  ' Cambiado a columna H (8)
        
        ' Crear CheckBoxes para cada celda en el rango
        For j = LBound(labels) To UBound(labels)
            With ws.CheckBoxes.Add(leftPos + j * checkBoxWidth, topPos, 20, 15) ' Ajustar el espaciado horizontal
                .Caption = labels(j)  ' Texto al lado del checkbox
                .Value = xlOff
                
                ' Establecer la celda de enlace en la hoja Datos
                .LinkedCell = wsLinks.Cells(i - 1, j + 1).Address(External:=True)
            End With
        Next j
    Next i
    
    ' Convertir los valores VERDADERO/FALSO en 1/0 en las celdas H1:L10
    For i = 1 To 10
        For j = 1 To 5
            wsLinks.Cells(i, j + 5).Formula = "=IF(" & wsLinks.Cells(i, j).Address & ",1,0)"
        Next j
        
        ' Concatenar los valores en la columna K (mantener concatenación en K)
        wsLinks.Cells(i, 11).Formula = "=" & wsLinks.Cells(i, 6).Address & "&" & wsLinks.Cells(i, 7).Address & "&" & wsLinks.Cells(i, 8).Address & "&" & wsLinks.Cells(i, 9).Address & "&" & wsLinks.Cells(i, 10).Address
    Next i
End Sub

'''

codigo_radiobuttons = '''
Sub CrearRadioButtonsRango()
    Dim ws As Worksheet
    Set ws = ActiveSheet
    
    ' Ruta del archivo de configuración
    Dim rutaConfig As String
    rutaConfig = "C:/Users/joaquin.rodriguezm/Desktop/config.txt"
    
    ' Leer archivo de configuración
    Dim config As Object
    Set config = CreateObject("Scripting.Dictionary")
    Call LeerArchivoConfiguracion(rutaConfig, config)
    
    ' Recorrer todos los rangos definidos en el archivo de configuración
    Dim clave As Variant
    Dim rangoCounter As Integer
    rangoCounter = 1

    For Each clave In config.Keys
        If InStr(clave, "rango") > 0 Then
            Dim rango As String
            rango = config(clave)
            
            ' Crear una nueva hoja para cada rango
            Dim wsDatos As Worksheet
            Set wsDatos = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
            wsDatos.Name = "Vinculaciones_Rango" & rangoCounter ' Nombre único para cada hoja
            rangoCounter = rangoCounter + 1
            
            ' Obtener las palabras y el número de botones para este rango
            Dim palabra1 As String, palabra2 As String, palabra3 As String
            Dim btnPalabra1 As Integer, btnPalabra2 As Integer, btnPalabra3 As Integer
            Dim nivelesPalabra1 As String, nivelesPalabra2 As String, nivelesPalabra3 As String
            
            ' Leer las palabras y el número de botones desde la configuración
            palabra1 = config("palabra" & rangoCounter - 1 & "1")
            palabra2 = config("palabra" & rangoCounter - 1 & "2")
            palabra3 = config("palabra" & rangoCounter - 1 & "3")
            btnPalabra1 = config("btnPalabra" & rangoCounter - 1 & "1")
            btnPalabra2 = config("btnPalabra" & rangoCounter - 1 & "2")
            btnPalabra3 = config("btnPalabra" & rangoCounter - 1 & "3")
            nivelesPalabra1 = config("nivelesPalabra" & rangoCounter - 1 & "1")
            nivelesPalabra2 = config("nivelesPalabra" & rangoCounter - 1 & "2")
            nivelesPalabra3 = config("nivelesPalabra" & rangoCounter - 1 & "3")
            
            ' Aplicar la lógica de creación de controles para cada rango
            Dim celda As Range, linkedCell As Range
            For Each celda In ws.Range(rango)

                celda.ColumnWidth = 53

                With celda.Borders
                    .LineStyle = xlContinuous
                    .Weight = xlThin
                End With

                ' Vincular a celdas en la nueva hoja creada para este rango
                Set linkedCell = wsDatos.Cells(celda.Row, 2) ' Usar la columna B en la nueva hoja
                
                ' Colocar texto en la celda
                celda.Value = Chr(10) & Chr(10) & palabra1 & Chr(10) & "" & Chr(10) & palabra2 & Chr(10) & "" & Chr(10) & palabra3 & Chr(10) & ""
                
                ' Ajustes de posicionamiento
                Dim top_offset As Integer
                Dim left_offset As Integer
                top_offset = celda.Top + 15
                left_offset = celda.Left + 60
                
                ' Posiciones ajustadas para los controles
                Dim posicionesY(1 To 3) As Integer
                posicionesY(1) = top_offset + 13    ' Ajuste para Palabra 1 (Velocidad)
                posicionesY(2) = top_offset + 43    ' Ajuste para Palabra 2 (Precisión)
                posicionesY(3) = top_offset + 70    ' Ajuste para Palabra 3 (Expresión)

                ' Crear Checkbox con nombre único
                Dim checkBoxP As Object
                Set checkBoxP = ws.CheckBoxes.Add(left_offset + 62, posicionesY(2) - 50, 70, 15)
                checkBoxP.Caption = "NO APLICA"
                checkBoxP.Name = "CheckBox" & celda.Address(False, False) ' Nombre único basado en la celda (e.g. CheckBoxH2)
                checkBoxP.OnAction = "BloquearOpcionesPorCheckbox"
                
                ' Crear GroupBox y OptionButtons para cada palabra
                Call CrearRadioButtons(ws, linkedCell, left_offset, posicionesY(1), palabra1, btnPalabra1, nivelesPalabra1, 0)
                Call CrearRadioButtons(ws, linkedCell, left_offset, posicionesY(2), palabra2, btnPalabra2, nivelesPalabra2, 1)
                Call CrearRadioButtons(ws, linkedCell, left_offset, posicionesY(3), palabra3, btnPalabra3, nivelesPalabra3, 2)
            Next celda
        End If
    Next clave
End Sub

' Subrutina para crear RadioButtons
Sub CrearRadioButtons(ws As Worksheet, linkedCell As Range, left_offset As Integer, top_offset As Integer, palabra As String, numBotones As Integer, niveles As String, offsetCol As Integer)
    Dim groupBox As Object
    Set groupBox = ws.GroupBoxes.Add(left_offset - 10, top_offset - 10, 250, 50)
    groupBox.Caption = palabra
    groupBox.Visible = False ' Ocultar GroupBox

    Dim j As Integer

    For j = 1 To numBotones
        Dim optButton As Object
        Set optButton = ws.OptionButtons.Add(left_offset + ((j - 1) * 60), top_offset, 60, 15)

        ' Asignar el nivel (letra) con el prefijo "NIVEL "
        If j <= Len(niveles) Then
            optButton.Caption = "NIVEL " & Mid(niveles, j, 1) ' Obtener el carácter j de la cadena de niveles
        Else
            optButton.Caption = "NIVEL " & j ' Asignar un nombre genérico si no hay más niveles
        End If
        
        ' Ajustar el offsetCol para usar las columnas C, D y E 
        optButton.LinkedCell = linkedCell.Offset(0, offsetCol + 1).Address(External:=True)
    Next j
End Sub

' Función para leer el archivo de configuración
Sub LeerArchivoConfiguracion(rutaConfig As String, ByRef config As Object)
    Dim fileNum As Integer
    Dim line As String
    fileNum = FreeFile

    Open rutaConfig For Input As fileNum
    Do Until EOF(fileNum)
        Line Input #fileNum, line
        If InStr(line, "=") > 0 Then
            Dim clave As String, valor As String
            clave = Trim(Left(line, InStr(line, "=") - 1))
            valor = Trim(Mid(line, InStr(line, "=") + 1))
            config(clave) = valor
        End If
    Loop
    Close fileNum
End Sub

'''

codigo_bloqueo = '''
Sub BloquearOpcionesPorCheckbox()
    Dim wsPreguntas As Worksheet
    Dim wsRespuestas As Worksheet
    Dim fila As Integer
    Dim col As Integer
    Dim checkEstado As Boolean
    Dim checkBoxName As String
    Dim celda As Range
    Dim optButton As Object
    Dim config As Object
    Dim rango As String
    Dim clave As Variant
    Dim rangoCounter As Integer
    Dim wsVinculaciones As Worksheet

    Application.ScreenUpdating = False

    Set wsPreguntas = ThisWorkbook.Sheets("Preguntas")
    Set wsRespuestas = ThisWorkbook.Sheets("Respuestas")

    ' Leer archivo de configuración
    Set config = CreateObject("Scripting.Dictionary")
    Call LeerArchivoConfiguracion("C:/Users/joaquin.rodriguezm/Desktop/config.txt", config)

    ' Iterar sobre los rangos definidos en el archivo de configuración
    For Each clave In config.Keys
        If InStr(clave, "rango") > 0 Then
            rango = config(clave)
            rangoCounter = Mid(clave, 6) ' Obtener el número del rango (ej: "1", "2", etc.)
            
            ' Establecer la hoja correspondiente al rango
            On Error Resume Next
            Set wsVinculaciones = ThisWorkbook.Sheets("Vinculaciones_Rango" & rangoCounter)
            On Error GoTo 0
            
            ' Asegurarse de que la hoja de vinculaciones existe
            If wsVinculaciones Is Nothing Then
                MsgBox "La hoja 'Vinculaciones_Rango" & rangoCounter & "' no existe.", vbExclamation
                Exit Sub
            End If
            
            For Each celda In wsPreguntas.Range(rango)
                fila = celda.Row
                col = celda.Column  ' Obtenemos la columna de la celda actual
                
                ' Crear el nombre del checkbox basado en la celda actual
                checkBoxName = "CheckBox" & celda.Address(False, False)

                ' Verificar el estado del checkbox correspondiente a esta celda
                On Error Resume Next ' Manejo de errores para checkboxes
                checkEstado = (wsPreguntas.CheckBoxes(checkBoxName).Value = 1)
                On Error GoTo 0

                ' Recorrer todos los OptionButtons para esta fila
                For Each optButton In wsPreguntas.OptionButtons
                    ' Verificar si el OptionButton está en la misma fila y columna
                    If optButton.TopLeftCell.Row = fila And optButton.TopLeftCell.Column = celda.Column Then
                        ' Si el checkbox está marcado, desmarcar y deshabilitar solo los OptionButtons en la misma fila
                        If checkEstado Then
                            optButton.Value = xlOff   ' Desmarcar el botón
                            optButton.Enabled = False  ' Deshabilitar el botón
                        Else
                            optButton.Enabled = True   ' Habilitar el botón si el checkbox no está marcado
                        End If
                    End If
                Next optButton

                ' Si el checkbox está marcado, escribir "NO APLICA" en la misma celda de la hoja Respuestas
                If checkEstado Then
                    wsRespuestas.Cells(fila, col).Value = "NO APLICA;NO APLICA;NO APLICA"
                Else
                    ' Aquí se coloca la fórmula que vincula a la celda correspondiente en Vinculaciones
                    wsRespuestas.Cells(fila, col).Formula = "='Vinculaciones_Rango" & rangoCounter & "'!" & wsVinculaciones.Cells(fila, 2).Address
                End If
            Next celda
        End If
    Next clave

    Application.ScreenUpdating = True
End Sub

'''

codigo_formulas = '''
Sub InsertarFormulasEnRango()
    Dim ws As Worksheet
    Dim config As Object
    Dim niveles As String
    Dim formula As String
    Dim i As Integer
    Dim j As Integer
    Dim colRef As String
    Dim palabraClave As String
    Dim rangoCounter As Integer

    ' Leer archivo de configuración
    Set config = CreateObject("Scripting.Dictionary")
    Call LeerArchivoConfiguracion("C:/Users/joaquin.rodriguezm/Desktop/config.txt", config)

    ' Iterar sobre los rangos definidos en el archivo de configuración
    rangoCounter = 1
    Do While config.Exists("rango" & rangoCounter)
        ' Establecer la hoja correspondiente al rango
        On Error Resume Next
        Set ws = ThisWorkbook.Sheets("Vinculaciones_Rango" & rangoCounter)
        On Error GoTo 0

        ' Comprobar si la hoja existe
        If Not ws Is Nothing Then
            ' Crear y establecer la fórmula celda por celda en el rango correspondiente
            For i = 2 To 11  ' Fila 2 a 11
                For j = 1 To 3  ' Para las palabras en columnas C (1), D (2), E (3)
                    ' Determinar la palabra clave según el rango y la posición
                    palabraClave = "nivelesPalabra" & rangoCounter & j
                    
                    ' Establecer las columnas de referencia
                    colRef = Chr(67 + j - 1) ' C, D, E (67 es la letra ASCII de 'C')

                    ' Verificar que la clave de niveles existe en el archivo de configuración
                    If config.Exists(palabraClave) Then
                        niveles = config(palabraClave)

                        ' Construir la fórmula de forma dinámica según los niveles
                        Select Case Len(niveles)
                            Case 2
                                formula = "=IF(" & colRef & i & "=1, ""NIVEL " & Mid(niveles, 1, 1) & """, IF(" & colRef & i & "=2, ""NIVEL " & Mid(niveles, 2, 1) & """, """"))"
                            Case 3
                                formula = "=IF(" & colRef & i & "=1, ""NIVEL " & Mid(niveles, 1, 1) & """, IF(" & colRef & i & "=2, ""NIVEL " & Mid(niveles, 2, 1) & """, IF(" & colRef & i & "=3, ""NIVEL " & Mid(niveles, 3, 1) & """, """")))"
                            Case 4
                                formula = "=IF(" & colRef & i & "=1, ""NIVEL " & Mid(niveles, 1, 1) & """, IF(" & colRef & i & "=2, ""NIVEL " & Mid(niveles, 2, 1) & """, IF(" & colRef & i & "=3, ""NIVEL " & Mid(niveles, 3, 1) & """, IF(" & colRef & i & "=4, ""NIVEL " & Mid(niveles, 4, 1) & """, """"))))"
                            ' Agregar más casos si es necesario
                        End Select

                        ' Insertar la fórmula celda por celda en la celda correspondiente, comenzando desde la columna F
                        ws.Cells(i, 6 + j - 1).Formula = formula ' Se ajusta a 6 + j para que empiece en F
                    Else
                        ' Si no existe la clave, podrías querer establecer un valor predeterminado o dejar la celda en blanco
                        ws.Cells(i, 6 + j).Value = "" ' O bien, puedes establecer un valor predeterminado aquí
                    End If
                Next j

                Dim concatFormula As String

                ' Crear fórmula para concatenar solo si las celdas no están vacías
                concatFormula = "=IF(F" & i & "<>"""", F" & i & ", """") & IF(AND(F" & i & "<>"""", G" & i & "<>""""), "";"" , """") & " & _
                "IF(G" & i & "<>"""", G" & i & ", """") & IF(AND(G" & i & "<>"""", H" & i & "<>""""), "";"" , """") & " & _
                "IF(H" & i & "<>"""", H" & i & ", """")"


                ' Insertar la fórmula concatenada en la columna B
                ws.Cells(i, 2).formula = concatFormula

                ' Ajustar el ancho de la columna B para que se ajusten todos los valores
                ws.Columns("B").ColumnWidth = 22

            Next i
        End If

        rangoCounter = rangoCounter + 1
    Loop
End Sub


'''

codigo_evento = '''
Private Sub Workbook_Open()
    AddCheckBoxes
    CrearRadioButtonsRango
    InsertarFormulasEnRango
    BloquearOpcionesPorCheckbox
End Sub
'''
# Agregar la macro al módulo de código
try:
    # Agregar un nuevo módulo
    modulo = libro.VBProject.VBComponents.Add(1)  # 1 para módulo estándar
    modulo.Name = "ModuloCheckboxes"  # Nombre del módulo
    modulo.CodeModule.AddFromString(codigo_checkboxes)  # Agregar el código de la macro

    modulo = libro.VBProject.VBComponents.Add(1)  # 1 para módulo estándar
    modulo.Name = "ModuloRadioButtons"  # Nombre del módulo
    modulo.CodeModule.AddFromString(codigo_radiobuttons)  # Agregar el código de la macro

    modulo = libro.VBProject.VBComponents.Add(1)  # 1 para módulo estándar
    modulo.Name = "ModuloBloqueo"  # Nombre del módulo
    modulo.CodeModule.AddFromString(codigo_bloqueo)  # Agregar el código de la macro

    modulo = libro.VBProject.VBComponents.Add(1)  # 1 para módulo estándar
    modulo.Name = "ModuloFormulas"  # Nombre del módulo
    modulo.CodeModule.AddFromString(codigo_formulas)  # Agregar el código de la macro

    ThisWorkbook = libro.VBProject.VBComponents("ThisWorkbook")
    ThisWorkbook.CodeModule.AddFromString(codigo_evento)


    # Guardar el libro
    libro.Save()
finally:
    # Cerrar el libro y Excel
    libro.Close(SaveChanges=True)
    excel.Quit()